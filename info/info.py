import xlwt
import xlrd
import numpy as np
from openpyxl import load_workbook
import random
def ge_excel():
    wb = xlwt.Workbook()
    # 添加一个表
    ws = wb.add_sheet('test')
    style = xlwt.XFStyle()
    style.num_format_str = '0.0'
    i = 0
    for k in range(5):
        i = 0
        for j in range(26):
            # 3个参数分别为行号，列号，和内容
            # 需要注意的是行号和列号都是从0开始的
            ws.write(i, 0, random.randint(361,370)/10)
            ws.write(i, 1, random.randint(361,370)/10)
            ws.write(i, 2, random.randint(361,370)/10)
            i = i + 1
        wb.save('D:\\桌面常用文件\\寒假信息统计\\返校\\每日信息\\test'+str(k)+'.xlsx')
def ge_excel_super():
    list = []
    x = 1000
    for j in range(1):
        x = x + 1
        list.append("{:.2f}".format(x/100))
    print(list)
    workbook = load_workbook('D:\\桌面常用文件\\寒假信息统计\\返校\\每日信息\\计研1903-9.11.xlsx')
    sheet = workbook['Sheet1']
    for j in list:
        for i in range(4,31):
            if i < 30:
                sheet.cell(i,2).value = j
                sheet.cell(i,7).value = random.randint(361,370)/10
                sheet.cell(i, 8).value = random.randint(361, 370) / 10
                sheet.cell(i, 9).value = random.randint(361, 370) / 10
            else:
                sheet.cell(i, 2).value = j
        workbook.save('D:\\桌面常用文件\\寒假信息统计\\返校\\每日信息\\计研1903-'+j+'.xlsx')
ge_excel_super()